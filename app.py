from flask import Flask, render_template, request, send_file
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from werkzeug.utils import secure_filename
import pandas as pd

app = Flask(__name__)

# Set the upload folder and allowed file extensions
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def process_excel(file_path):
    # # Load the uploaded Excel file
    # wb = load_workbook(file_path)
    # ws = wb.active

    # # Create a new workbook to save the processed data
    # new_wb = Workbook()
    # new_ws = new_wb.active

    # # Example processing: Copy the first sheet into the new file
    # for i in ws.iter_cols:
    #     print(i)
    # for row in ws.iter_rows(values_only=True):

    #     new_ws.append(row)
  # Load data from Excel into a DataFrame
    df = pd.read_excel(file_path, engine='openpyxl')

    df = df[df['देयकाची एकूण रक्कम'].notna() & (df['देयकाची एकूण रक्कम'] > 0)]


    # Ensure the 'date' column is a datetime type
    df['date'] = pd.to_datetime(df['दिनांक'])

    # Extract Year-Month from the date column
    df['year_month'] = df['date'].dt.to_period('M')

    # Group by person and year_month, then sum the tax
    result = df.groupby(['जीएसटी क्रमांक', 'year_month'])['देयकाची एकूण रक्कम'].sum().reset_index()
    # Display the DataFrame
    # print(df["कंत्राटदाराचे नाव"])
    # for index, row in df.iterrows():
    #     print(f"Index: {index}, A: {row['कंत्राटदाराचे नाव']}, B: {row['देयकाची एकूण रक्कम']}")

    # with pd.ExcelWriter('tax_summary.xlsx') as writer:
    #     result.to_excel(writer, sheet_name='Tax_Summary', index=False)

    pivoted_result = result.pivot(index='जीएसटी क्रमांक', columns='year_month', values='देयकाची एकूण रक्कम')

    # Reset the column names to show in proper format
    pivoted_result.columns = pivoted_result.columns.astype(str)

   
    # You can add more processing logic here
    output_file = os.path.join(app.config['UPLOAD_FOLDER'], 'processed_tipni_file.xlsx')

    # Save to Excel
    pivoted_result.to_excel(output_file,engine='openpyxl')

    return output_file

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part', 400
        file = request.files['file']
        if file.filename == '':
            return 'No selected file', 400
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Process the Excel file
            processed_file = process_excel(file_path)
            
            # Return the processed file for download
            return send_file(processed_file, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
