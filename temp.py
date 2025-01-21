
import os
from openpyxl import load_workbook
from openpyxl import Workbook

import pandas as pd


def process_excel(file_path):
    # Load the uploaded Excel file
    # wb = load_workbook(file_path)
    # ws = wb.active

    # # Create a new workbook to save the processed data
    # new_wb = Workbook()
    # new_ws = new_wb.active

    # # Example processing: Copy the first sheet into the new file

    # print(ws.columns)
    # count=0
    # for row in ws.iter_rows(values_only=True):
    #     if count<5:
    #         print(row)
    #     count+=1
        
    # for row in ws.iter_rows(values_only=True):
    #     new_ws.append(row)

    # You can add more processing logic here
    # output_file = os.path.join(app.config['UPLOAD_FOLDER'], 'processed_file.xlsx')
    # new_wb.save(output_file)

    # return output_file
    # Load data from Excel into a DataFrame
    df = pd.read_excel(file_path, engine='openpyxl',skiprows=3)

    # Ensure the 'date' column is a datetime type
    df['date'] = pd.to_datetime(df['दिनांक'])

    # Extract Year-Month from the date column
    df['year_month'] = df['date'].dt.to_period('M')

    # Group by person and year_month, then sum the tax
    result = df.groupby(['कंत्राटदाराचे नाव', 'year_month'])['देयकाची एकूण रक्कम'].sum().reset_index()
    # Display the DataFrame
    # print(df["कंत्राटदाराचे नाव"])
    # for index, row in df.iterrows():
    #     print(f"Index: {index}, A: {row['कंत्राटदाराचे नाव']}, B: {row['देयकाची एकूण रक्कम']}")

    # with pd.ExcelWriter('tax_summary.xlsx') as writer:
    #     result.to_excel(writer, sheet_name='Tax_Summary', index=False)

    pivoted_result = result.pivot(index='कंत्राटदाराचे नाव', columns='year_month', values='देयकाची एकूण रक्कम')

    # Reset the column names to show in proper format
    pivoted_result.columns = pivoted_result.columns.astype(str)

    # Save to Excel
    pivoted_result.to_excel('tipni_summary_per_month.xlsx',engine='openpyxl')

    print(pivoted_result)
    # result.to_excel('tax_summary.xlsx', index=False,engine='openpyxl')


process_excel("C:/Users/admin/Downloads/tipniData2024.xlsx")








